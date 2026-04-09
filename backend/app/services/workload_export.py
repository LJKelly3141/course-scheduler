"""Export instructor workload to XLSX matching department report format."""
from __future__ import annotations

import io
import re
from typing import Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.term import Term
from app.services.workload import compute_instructor_workload


# Column indices (1-based)
COL_LAST = 1
COL_FIRST = 2
COL_STATUS = 3
COL_DEPT = 4
COL_COURSE = 5
COL_SECTION = 6
COL_INSTR_TYPE = 7   # Lec, Lab or Fld
COL_CLASS_NAME = 8
COL_ENROLL = 9
COL_ACTUAL_CR = 10
COL_EQUIV_CR = 11
COL_SCH = 12
COL_TOTAL_LOAD = 13
COL_REASSIGN = 14
COL_FORMS = 15

HEADERS = [
    "Last Name",
    "First Name",
    "Status",
    "Dept Code",
    "Course #",
    "Section #",
    "Lec, Lab or Fld",
    "Class Name",
    "Enrollment",
    "Actual Credits",
    "Equivalent Credits",
    "SCH",
    "TOTAL LOAD for Semester",
    "Reassignment/Suggestion",
    "Forms completed?",
]

COL_WIDTHS = [14, 12, 10, 10, 10, 10, 14, 28, 12, 14, 16, 8, 20, 24, 16]

# Styles
_header_font = Font(bold=True, size=11)
_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_bold_font = Font(bold=True)
_thin_border = Border(bottom=Side(style="thin"))
_wrap_align = Alignment(wrap_text=True, vertical="top")
_overload_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

_STATUS_MAP = {
    "Faculty": "F",
    "faculty": "F",
    "IAS": "IAS",
    "ias": "IAS",
    "Adjunct": "ADJ",
    "adjunct": "ADJ",
    "NIAS": "NIAS",
    "nias": "NIAS",
}

_ADJ_TYPE_LABELS = {
    "research_release": "Research Reassignment",
    "admin_release": "Admin Reassignment",
    "course_release": "Course Reassignment",
    "adhoc": "ADHOC",
    "overload": "Overload",
    "other": "Other",
}


def _status_abbrev(instructor_type: str) -> str:
    """Convert instructor_type to short status code for the report."""
    if not instructor_type:
        return ""
    return _STATUS_MAP.get(instructor_type, instructor_type.upper())


def _safe_sheet_name(name: str) -> str:
    """Sanitize a string for use as an Excel sheet name (max 31 chars)."""
    name = re.sub(r'[\\/*?\[\]:]', '', name)
    return name[:31]


def export_workload_xlsx(db: Session, term_id: int) -> Tuple[bytes, str]:
    """Generate an Excel workbook for instructor workload.

    Returns (xlsx_bytes, suggested_filename).
    """
    data = compute_instructor_workload(db, term_id)

    # Load term for sheet name and filename
    term = db.query(Term).filter(Term.id == term_id).first()
    term_label = term.name if term else f"Term {term_id}"
    sheet_name = _safe_sheet_name(term_label)
    safe_filename = re.sub(r'[^\w\s-]', '', term_label).strip().replace(' ', '_')
    filename = f"faculty_load_{safe_filename}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write headers
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = _wrap_align

    # Column widths
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 2

    for inst in data["instructors"]:
        first_data_row = row
        status = _status_abbrev(inst["instructor_type"])

        # --- Section rows ---
        for sect in inst["sections"]:
            ws.cell(row=row, column=COL_LAST, value=inst["last_name"])
            ws.cell(row=row, column=COL_FIRST, value=inst["first_name"])
            ws.cell(row=row, column=COL_STATUS, value=status)
            ws.cell(row=row, column=COL_DEPT, value=sect["department_code"])
            ws.cell(row=row, column=COL_COURSE, value=sect["course_number"])
            ws.cell(row=row, column=COL_SECTION, value=sect["section_number"])
            ws.cell(row=row, column=COL_INSTR_TYPE, value=sect.get("instruction_type", "LEC"))
            ws.cell(row=row, column=COL_CLASS_NAME, value=sect["title"])
            ws.cell(row=row, column=COL_ENROLL, value=sect["enrollment_cap"])
            ws.cell(row=row, column=COL_ACTUAL_CR, value=sect["actual_credits"])
            ws.cell(row=row, column=COL_EQUIV_CR, value=sect["equivalent_credits"])
            ws.cell(row=row, column=COL_SCH, value=sect["sch"])
            row += 1

        # --- Adjustment / reassignment rows ---
        for adj in inst["adjustments"]:
            ws.cell(row=row, column=COL_LAST, value=inst["last_name"])
            ws.cell(row=row, column=COL_FIRST, value=inst["first_name"])
            ws.cell(row=row, column=COL_STATUS, value=status)
            ws.cell(row=row, column=COL_DEPT, value=inst["department"])
            ws.cell(row=row, column=COL_COURSE, value="Release")
            ws.cell(row=row, column=COL_CLASS_NAME, value=adj["description"])
            ws.cell(row=row, column=COL_EQUIV_CR, value=adj["equivalent_credits"])
            ws.cell(row=row, column=COL_REASSIGN, value=_ADJ_TYPE_LABELS.get(
                adj["adjustment_type"], adj["adjustment_type"]
            ))
            row += 1

        # If instructor has no sections and no adjustments, write a placeholder
        if not inst["sections"] and not inst["adjustments"]:
            ws.cell(row=row, column=COL_LAST, value=inst["last_name"])
            ws.cell(row=row, column=COL_FIRST, value=inst["first_name"])
            ws.cell(row=row, column=COL_STATUS, value=status)
            ws.cell(row=row, column=COL_DEPT, value=inst["department"])
            row += 1

        last_data_row = row - 1

        # --- Subtotal row with SUM formulas ---
        for col_idx in range(1, len(HEADERS) + 1):
            ws.cell(row=row, column=col_idx).border = _thin_border

        ws.cell(row=row, column=COL_LAST, value=inst["last_name"]).font = _bold_font
        ws.cell(row=row, column=COL_FIRST, value=inst["first_name"]).font = _bold_font

        # SUM formulas for credits columns
        if first_data_row <= last_data_row:
            actual_col_letter = get_column_letter(COL_ACTUAL_CR)
            equiv_col_letter = get_column_letter(COL_EQUIV_CR)
            sch_col_letter = get_column_letter(COL_SCH)

            actual_cell = ws.cell(
                row=row, column=COL_ACTUAL_CR,
                value=f"=SUM({actual_col_letter}{first_data_row}:{actual_col_letter}{last_data_row})"
            )
            actual_cell.font = _bold_font

            equiv_cell = ws.cell(
                row=row, column=COL_EQUIV_CR,
                value=f"=SUM({equiv_col_letter}{first_data_row}:{equiv_col_letter}{last_data_row})"
            )
            equiv_cell.font = _bold_font

            sch_cell = ws.cell(
                row=row, column=COL_SCH,
                value=f"=SUM({sch_col_letter}{first_data_row}:{sch_col_letter}{last_data_row})"
            )
            sch_cell.font = _bold_font
        else:
            actual_cell = ws.cell(row=row, column=COL_ACTUAL_CR, value=0)
            actual_cell.font = _bold_font
            equiv_cell = ws.cell(row=row, column=COL_EQUIV_CR, value=0)
            equiv_cell.font = _bold_font
            sch_cell = ws.cell(row=row, column=COL_SCH, value=0)
            sch_cell.font = _bold_font

        # Total Load = SUM of equivalent credits (includes adjustment rows)
        if first_data_row <= last_data_row:
            load_cell = ws.cell(
                row=row, column=COL_TOTAL_LOAD,
                value=f"=SUM({equiv_col_letter}{first_data_row}:{equiv_col_letter}{last_data_row})"
            )
        else:
            load_cell = ws.cell(row=row, column=COL_TOTAL_LOAD, value=0)
        load_cell.font = _bold_font

        if inst["is_overloaded"]:
            load_cell.fill = _overload_fill

        row += 1
        row += 1  # blank separator row

    # --- Unassigned sections ---
    if data["unassigned_sections"]:
        ws.cell(row=row, column=COL_LAST, value="NEEDS TO BE ASSIGNED").font = Font(bold=True, size=12)
        row += 1

        for sect in data["unassigned_sections"]:
            ws.cell(row=row, column=COL_DEPT, value=sect["department_code"])
            ws.cell(row=row, column=COL_COURSE, value=sect["course_number"])
            ws.cell(row=row, column=COL_SECTION, value=sect["section_number"])
            ws.cell(row=row, column=COL_INSTR_TYPE, value=sect.get("instruction_type", "LEC"))
            ws.cell(row=row, column=COL_CLASS_NAME, value=sect["title"])
            ws.cell(row=row, column=COL_ENROLL, value=sect["enrollment_cap"])
            ws.cell(row=row, column=COL_ACTUAL_CR, value=sect["actual_credits"])
            ws.cell(row=row, column=COL_EQUIV_CR, value=sect["equivalent_credits"])
            ws.cell(row=row, column=COL_SCH, value=sect["sch"])
            row += 1

    # Write to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue(), filename
