"""
Generic XLSX reader with smart/fuzzy column detection.

Reads XLSX files (via openpyxl) and maps spreadsheet column headers to
canonical column names using a three-phase approach:
1. Exact match after normalization (includes known aliases)
2. Word-containment match (e.g. "Title" matches "course title")
3. Fuzzy match via difflib.SequenceMatcher for remaining headers
"""
from __future__ import annotations

import csv
import difflib
import io
import re
from typing import Dict, List, Optional, Tuple

import openpyxl


# ---------------------------------------------------------------------------
# Column alias maps per import type
# ---------------------------------------------------------------------------
# Keys = canonical column names (what validation functions expect).
# Values = lists of known alternative headers (will be normalized for comparison).

ROOM_ALIASES: Dict[str, List[str]] = {
    "building_name": [
        "building name", "building", "bldg name", "bldg",
    ],
    "building_abbreviation": [
        "building abbreviation", "abbreviation", "bldg abbreviation",
        "bldg abbr", "bldg code", "building code", "building abbr",
        "abbr", "code",
    ],
    "room_number": [
        "room number", "room", "room num", "room no", "room #", "number",
        "rm", "rm #", "rm no",
    ],
    "capacity": [
        "capacity", "cap", "seats", "max capacity", "seating capacity",
        "room capacity", "size",
    ],
}

INSTRUCTOR_ALIASES: Dict[str, List[str]] = {
    "name": [
        "name", "instructor name", "instructor", "full name",
        "faculty name", "faculty", "prof", "professor",
    ],
    "email": [
        "email", "email address", "e-mail", "e mail",
        "mail", "contact", "contact email",
    ],
    "department": [
        "department", "dept", "department code", "dept code",
        "dept name", "department name",
    ],
    "modality_constraint": [
        "modality constraint", "modality", "teaching modality",
        "modality preference", "delivery mode", "instruction mode",
    ],
    "max_credits": [
        "max credits", "maximum credits", "credit limit", "max credit load",
        "credit load", "max load",
    ],
}

COURSE_ALIASES: Dict[str, List[str]] = {
    "department_code": [
        "department code", "dept code", "department", "dept", "subject",
        "subject code", "subj", "prefix",
    ],
    "course_number": [
        "course number", "course num", "course no", "course #", "number",
        "catalog number", "catalog no", "catalog #", "num", "no",
    ],
    "title": [
        "title", "course title", "course name", "name", "description",
        "course description",
    ],
    "credits": [
        "credits", "credit hours", "credit", "hrs", "hours",
        "cr", "cr hrs",
    ],
}

SCHEDULE_ALIASES: Dict[str, List[str]] = {
    "Class": [
        "class", "crn", "class number", "class #", "class no",
        "class nbr", "class num", "reference number", "ref",
    ],
    "Course": [
        "course", "course name", "subject course",
        "subject", "subject/course", "course/section",
    ],
    "Title": [
        "title", "course title", "course description",
        "name", "description",
    ],
    "Section": [
        "section", "section number", "sec", "section #",
        "section no", "sec #", "sec no", "section num",
    ],
    "Days & Times": [
        "days & times", "days and times", "days times", "meeting times",
        "days/times", "schedule", "meeting pattern",
        "day/time", "day & time", "day and time",
        "time", "times", "days", "day",
        "meeting time", "meeting info",
    ],
    "Room": [
        "room", "location", "room location", "bldg/room", "building room",
        "bldg room", "bldg/rm", "building/room", "where",
        "room #", "room number", "rm",
    ],
    "Instructor": [
        "instructor", "instructor name", "faculty", "faculty name",
        "teacher", "professor", "prof",
        "assigned instructor", "primary instructor",
    ],
    "Meeting Dates": [
        "meeting dates", "dates", "term dates", "start end dates",
        "date range", "start/end", "start end", "start date",
    ],
    "Status": [
        "status", "class status", "enrollment status",
    ],
    "Enrollment": [
        "enrollment", "enrolled", "enrollment total", "enrl tot",
        "actual enrollment",
    ],
    "Capacity": [
        "capacity", "cap", "enrollment cap", "enrl cap",
        "max enrollment",
    ],
}


# ---------------------------------------------------------------------------
# Normalisation & matching
# ---------------------------------------------------------------------------

def _normalize(text: str) -> str:
    """Normalize a column header for comparison."""
    s = text.strip().lower()
    s = s.replace("_", " ").replace("-", " ")
    s = re.sub(r"\s+", " ", s)
    return s


def match_columns(
    actual_headers: List[str],
    alias_map: Dict[str, List[str]],
    fuzzy_threshold: float = 0.75,
) -> Tuple[Dict[str, str], List[str], List[str]]:
    """Match actual spreadsheet headers to canonical column names.

    Returns (column_mapping, warnings, errors):
      - column_mapping: Dict[actual_header -> canonical_name]
      - warnings: informational messages (fuzzy matches, ignored columns)
      - errors: blocking errors (not currently produced here)
    """
    column_mapping: Dict[str, str] = {}
    warnings: List[str] = []
    matched_canonical: set = set()

    # Build reverse lookup: normalized_alias -> canonical_name
    reverse_lookup: Dict[str, str] = {}
    for canonical, aliases in alias_map.items():
        reverse_lookup[_normalize(canonical)] = canonical
        for alias in aliases:
            reverse_lookup[_normalize(alias)] = canonical

    # Phase 1: exact match after normalisation
    unmatched_headers: List[str] = []
    for header in actual_headers:
        normalized = _normalize(header)
        if normalized in reverse_lookup:
            canonical = reverse_lookup[normalized]
            if canonical not in matched_canonical:
                column_mapping[header] = canonical
                matched_canonical.add(canonical)
            else:
                warnings.append(
                    f"Duplicate match: '{header}' also maps to "
                    f"'{canonical}' (already matched)"
                )
        else:
            unmatched_headers.append(header)

    # Phase 2: word-containment match — if the header appears as a whole word
    # within a multi-word alias (or vice versa), treat it as a match.
    # This catches short headers like "Title" matching "course title".
    unmatched_canonical = set(alias_map.keys()) - matched_canonical
    phase2_unmatched: List[str] = []

    for header in unmatched_headers:
        normalized = _normalize(header)
        norm_words = normalized.split()
        best_canonical: Optional[str] = None

        for canonical in unmatched_canonical:
            candidates = [_normalize(canonical)] + [
                _normalize(a) for a in alias_map[canonical]
            ]
            for candidate in candidates:
                cand_words = candidate.split()
                # Header words are a subset of alias words (or equal)
                if norm_words and all(w in cand_words for w in norm_words):
                    best_canonical = canonical
                    break
                # Alias words are a subset of header words (or equal)
                if cand_words and all(w in norm_words for w in cand_words):
                    best_canonical = canonical
                    break
            if best_canonical:
                break

        if best_canonical:
            column_mapping[header] = best_canonical
            matched_canonical.add(best_canonical)
            unmatched_canonical.discard(best_canonical)
            warnings.append(
                f"Matched column '{header}' -> '{best_canonical}' "
                f"(word match)"
            )
        else:
            phase2_unmatched.append(header)

    # Phase 3: fuzzy match remaining headers against unmatched canonical names
    still_unmatched: List[str] = []

    for header in phase2_unmatched:
        normalized = _normalize(header)
        best_score = 0.0
        best_canonical = None

        for canonical in unmatched_canonical:
            candidates = [_normalize(canonical)] + [
                _normalize(a) for a in alias_map[canonical]
            ]
            for candidate in candidates:
                score = difflib.SequenceMatcher(
                    None, normalized, candidate
                ).ratio()
                if score > best_score:
                    best_score = score
                    best_canonical = canonical

        if best_score >= fuzzy_threshold and best_canonical:
            column_mapping[header] = best_canonical
            matched_canonical.add(best_canonical)
            unmatched_canonical.discard(best_canonical)
            warnings.append(
                f"Fuzzy matched column '{header}' -> '{best_canonical}' "
                f"(score: {best_score:.0%})"
            )
        else:
            still_unmatched.append(header)

    for header in still_unmatched:
        if header.strip():
            warnings.append(f"Ignoring unrecognized column: '{header}'")

    return column_mapping, warnings, []


# ---------------------------------------------------------------------------
# File readers
# ---------------------------------------------------------------------------

def read_xlsx_headers(
    contents: bytes,
) -> Tuple[List[str], List[str]]:
    """Read just the headers from an XLSX file.

    Returns (headers, errors).
    """
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(contents), read_only=True, data_only=True
        )
    except Exception as exc:
        return [], [f"Could not read XLSX file: {exc}"]

    ws = wb.active
    if ws is None:
        wb.close()
        return [], ["XLSX file has no active sheet"]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        wb.close()
        return [], ["XLSX file is empty"]

    actual_headers = [str(c).strip() if c else "" for c in header]
    wb.close()
    return actual_headers, []


def read_xlsx_to_rows(
    contents: bytes,
    alias_map: Dict[str, List[str]],
    required_columns: Optional[List[str]] = None,
    column_mapping_override: Optional[Dict[str, str]] = None,
) -> Tuple[List[dict], List[str]]:
    """Read an XLSX file and return rows with canonicalized column names.

    Returns (rows, messages) where rows are dicts keyed by canonical names.
    All values are strings (stripped). Messages include warnings and errors.

    If column_mapping_override is provided, it is used instead of auto-detection.
    Keys are actual header names, values are canonical column names.
    """
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(contents), read_only=True, data_only=True
        )
    except Exception as exc:
        return [], [f"Could not read XLSX file: {exc}"]

    ws = wb.active
    if ws is None:
        wb.close()
        return [], ["XLSX file has no active sheet"]

    rows_iter = ws.iter_rows(values_only=True)

    # Read header row
    try:
        header = next(rows_iter)
    except StopIteration:
        wb.close()
        return [], ["XLSX file is empty"]

    actual_headers = [str(c).strip() if c else "" for c in header]

    # Match columns — use override if provided
    if column_mapping_override:
        column_mapping = column_mapping_override
        warnings: List[str] = []
        errors: List[str] = []
    else:
        column_mapping, warnings, errors = match_columns(actual_headers, alias_map)

    # Check required columns
    if required_columns is None:
        required_columns = list(alias_map.keys())

    matched_canonical = set(column_mapping.values())
    missing = [col for col in required_columns if col not in matched_canonical]
    if missing:
        hints = []
        for col in missing:
            aliases = alias_map.get(col, [])
            hint = f" (also accepts: {', '.join(aliases[:3])})" if aliases else ""
            hints.append(f"'{col}'{hint}")
        errors.append(f"Missing required columns: {', '.join(hints)}")
        wb.close()
        return [], errors + warnings

    # Parse data rows
    rows: List[dict] = []
    for row_values in rows_iter:
        if all(v is None or str(v).strip() == "" for v in row_values):
            continue

        row_dict: dict = {}
        for i, actual_header in enumerate(actual_headers):
            if actual_header in column_mapping and i < len(row_values):
                canonical = column_mapping[actual_header]
                value = row_values[i]
                row_dict[canonical] = str(value).strip() if value is not None else ""

        rows.append(row_dict)

    wb.close()
    all_messages = warnings + errors
    return rows, all_messages


def read_csv_to_rows(contents: bytes) -> List[dict]:
    """Read CSV bytes into a list of row dicts (canonical column names expected)."""
    text = contents.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [row for row in reader]
