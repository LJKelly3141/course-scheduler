from __future__ import annotations

import csv
import io
import json
from datetime import time

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session, joinedload

from app.models.meeting import Meeting
from app.models.section import Section
from app.models.room import Room
from app.models.building import Building
from app.models.instructor import Instructor
from app.models.course import Course
from app.models.term import Term
from app.models.term_session import TermSession


def _format_time(t) -> str:
    """Format a time object as HH:MM AM/PM."""
    if t is None:
        return ""
    if isinstance(t, time):
        return t.strftime("%-I:%M %p")
    return str(t)


def _parse_days(days_json: str) -> str:
    """Convert a JSON list like '["M","W","F"]' into a compact string like 'MWF'."""
    try:
        days = json.loads(days_json)
        if isinstance(days, list):
            return "".join(days)
    except (json.JSONDecodeError, TypeError):
        pass
    return days_json or ""


def _parse_days_list(days_json: str) -> list:
    """Parse days JSON into a list of day codes."""
    try:
        days = json.loads(days_json)
        if isinstance(days, list):
            return days
    except (json.JSONDecodeError, TypeError):
        pass
    return []


def _format_time_range(start, end) -> str:
    """Format start and end times as a single range string."""
    s = _format_time(start)
    e = _format_time(end)
    if s and e:
        return f"{s} - {e}"
    return s or e or ""


# ---------------------------------------------------------------------------
# CSV Export
# ---------------------------------------------------------------------------
def export_term_csv(db: Session, term_id: int) -> str:
    """Generate a CSV string containing all meetings for the given term."""
    term = db.query(Term).filter(Term.id == term_id).first()
    if not term:
        raise ValueError(f"Term {term_id} not found")

    meetings = (
        db.query(Meeting)
        .join(Section, Meeting.section_id == Section.id)
        .filter(Section.term_id == term_id)
        .options(
            joinedload(Meeting.section).joinedload(Section.course),
            joinedload(Meeting.section).joinedload(Section.term_session),
            joinedload(Meeting.room).joinedload(Room.building),
            joinedload(Meeting.instructor),
        )
        .all()
    )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Course Code",
        "Section",
        "Title",
        "Credits",
        "Session",
        "Start Date",
        "End Date",
        "Days",
        "Start Time",
        "End Time",
        "Room",
        "Building",
        "Instructor",
        "Enrollment Cap",
        "Modality",
    ])

    for m in meetings:
        section = m.section
        course = section.course if section else None
        room = m.room
        building = room.building if room else None
        instructor = m.instructor
        ts = section.term_session if section else None

        course_code = (
            f"{course.department_code} {course.course_number}" if course else ""
        )
        writer.writerow([
            course_code,
            section.section_number if section else "",
            course.title if course else "",
            course.credits if course else "",
            ts.name if ts else "",
            str(ts.start_date) if ts and ts.start_date else "",
            str(ts.end_date) if ts and ts.end_date else "",
            _parse_days(m.days_of_week),
            _format_time(m.start_time),
            _format_time(m.end_time),
            room.room_number if room else "",
            building.abbreviation if building else "",
            instructor.name if instructor else "",
            section.enrollment_cap if section else "",
            section.modality if section else "",
        ])

    return output.getvalue()


# ---------------------------------------------------------------------------
# XLSX Export (Dean's Office Schedule Table)
# ---------------------------------------------------------------------------
def export_term_xlsx(db: Session, term_id: int) -> bytes:
    """Generate an XLSX workbook containing all sections for the given term."""
    term = db.query(Term).filter(Term.id == term_id).first()
    if not term:
        raise ValueError(f"Term {term_id} not found")

    meetings = (
        db.query(Meeting)
        .join(Section, Meeting.section_id == Section.id)
        .filter(Section.term_id == term_id)
        .options(
            joinedload(Meeting.section).joinedload(Section.course),
            joinedload(Meeting.section).joinedload(Section.term_session),
            joinedload(Meeting.room).joinedload(Room.building),
            joinedload(Meeting.instructor),
        )
        .all()
    )

    # Also get sections with no meetings (online/unscheduled)
    sections_with_meetings = {m.section_id for m in meetings}
    all_sections = (
        db.query(Section)
        .filter(Section.term_id == term_id)
        .options(
            joinedload(Section.course),
            joinedload(Section.instructor),
            joinedload(Section.term_session),
        )
        .all()
    )
    unscheduled_sections = [
        s for s in all_sections if s.id not in sections_with_meetings
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = term.name

    # UWRF "Schedule Changes, Additions, Deletions" form columns
    headers = [
        "Action", "Dept Abbreviation", "Catalog Number", "Section",
        "Course Title", "Credits", "Session", "Start Date", "End Date",
        "Time Begin-End",
        "M", "T", "W", "R", "F", "S",
        "Instructor", "Building", "Room Number",
        "Lecture Hours", "Special Course Fee",
        "Maximum Class Size", "Instruction Mode", "Class Notes",
    ]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Map internal day codes to UWRF form day columns
    # Our model uses "Th" for Thursday; UWRF form uses "R"
    _DAY_COLS = ["M", "T", "W", "Th", "F", "S"]

    def _day_markers(days_json: str) -> tuple:
        """Return a tuple of 6 values (M, T, W, R, F, S) with 'X' or ''."""
        days = _parse_days_list(days_json)
        return tuple("X" if d in days else "" for d in _DAY_COLS)

    # Build rows: one per meeting, plus one per unscheduled section
    rows = []
    for m in meetings:
        section = m.section
        course = section.course if section else None
        room = m.room
        building = room.building if room else None
        instructor = m.instructor
        ts = section.term_session if section else None
        day_m, day_t, day_w, day_r, day_f, day_s = _day_markers(m.days_of_week)
        rows.append((
            "ADD",
            course.department_code if course else "",
            course.course_number if course else "",
            section.section_number if section else "",
            course.title if course else "",
            course.credits if course else "",
            ts.name if ts else "",
            str(ts.start_date) if ts and ts.start_date else "",
            str(ts.end_date) if ts and ts.end_date else "",
            _format_time_range(m.start_time, m.end_time),
            day_m, day_t, day_w, day_r, day_f, day_s,
            instructor.name if instructor else "",
            building.abbreviation if building else "",
            room.room_number if room else "",
            section.lecture_hours if section else "",
            section.special_course_fee if section else "",
            section.enrollment_cap if section else "",
            section.modality if section else "",
            section.notes if section and section.notes else ""
        ))

    for s in unscheduled_sections:
        course = s.course
        ts = s.term_session
        rows.append((
            "ADD",
            course.department_code if course else "",
            course.course_number if course else "",
            s.section_number,
            course.title if course else "",
            course.credits if course else "",
            ts.name if ts else "",
            str(ts.start_date) if ts and ts.start_date else "",
            str(ts.end_date) if ts and ts.end_date else "",
            "TBA",
            "", "", "", "", "", "",
            s.instructor.name if s.instructor else "",
            "On-Line" if s.modality and "online" in str(s.modality).lower() else "",
            "",
            s.lecture_hours if s.lecture_hours else "",
            s.special_course_fee if s.special_course_fee else "",
            s.enrollment_cap,
            s.modality if s.modality else "",
            s.notes if s.notes else ""
        ))

    # Sort by Dept Abbreviation, Catalog Number, then Section
    rows.sort(key=lambda r: (r[1], r[2], r[3]))

    for row_data in rows:
        ws.append(list(row_data))

    # Auto-fit column widths
    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Printable HTML Export
# ---------------------------------------------------------------------------
_CSS = """\
<style>
    * { margin: 0; padding: 0; box-sizing: border-box; }
    body { font-family: Arial, Helvetica, sans-serif; font-size: 11px; color: #333; padding: 20px; }
    h1 { font-size: 18px; margin-bottom: 4px; }
    h2 { font-size: 14px; margin-top: 18px; margin-bottom: 6px; border-bottom: 2px solid #333; padding-bottom: 2px; }
    table { border-collapse: collapse; width: 100%; margin-bottom: 16px; }
    th, td { border: 1px solid #999; padding: 3px 6px; text-align: left; vertical-align: top; }
    th { background: #e9e9e9; font-weight: bold; }
    .empty { color: #aaa; }
    @media print {
        body { padding: 0; }
        h2 { page-break-before: auto; }
        table { page-break-inside: auto; }
        tr { page-break-inside: avoid; }
    }
</style>
"""


def _build_meeting_map(meetings: list[Meeting]) -> dict:
    """Return a dict grouping meetings by various keys."""
    by_room: dict[str, list[Meeting]] = {}
    by_instructor: dict[str, list[Meeting]] = {}

    for m in meetings:
        # Room key
        if m.room and m.room.building:
            room_key = f"{m.room.building.abbreviation} {m.room.room_number}"
        elif m.room:
            room_key = m.room.room_number
        else:
            room_key = "Unassigned"
        by_room.setdefault(room_key, []).append(m)

        # Instructor key
        instr_key = m.instructor.name if m.instructor else "Unassigned"
        by_instructor.setdefault(instr_key, []).append(m)

    return {"by_room": by_room, "by_instructor": by_instructor}


def _meeting_cell(m: Meeting) -> str:
    """Render a compact HTML snippet for a meeting."""
    section = m.section
    course = section.course if section else None
    code = f"{course.department_code} {course.course_number}" if course else "?"
    sec = section.section_number if section else ""
    instr = m.instructor.name if m.instructor else ""
    room_label = ""
    if m.room:
        bldg = m.room.building.abbreviation if m.room.building else ""
        room_label = f"{bldg} {m.room.room_number}"
    return f"{code}-{sec}<br>{instr}<br>{room_label}"


def _time_slots() -> list[tuple[str, time, time]]:
    """Generate standard time slots from 8 AM to 6 PM in 1-hour blocks."""
    slots = []
    for hour in range(8, 18):
        start = time(hour, 0)
        end = time(hour + 1, 0)
        label = f"{start.strftime('%-I:%M %p')} - {end.strftime('%-I:%M %p')}"
        slots.append((label, start, end))
    return slots


def _meeting_overlaps_slot(m: Meeting, slot_start: time, slot_end: time) -> bool:
    """Check if a meeting overlaps with a given time slot."""
    m_start = m.start_time if isinstance(m.start_time, time) else time(0, 0)
    m_end = m.end_time if isinstance(m.end_time, time) else time(0, 0)
    return m_start < slot_end and m_end > slot_start


_DAYS = ["M", "T", "W", "Th", "F"]


def _meeting_on_day(m: Meeting, day: str) -> bool:
    """Check if a meeting occurs on the given day code."""
    try:
        days = json.loads(m.days_of_week)
        if isinstance(days, list):
            return day in days
    except (json.JSONDecodeError, TypeError):
        pass
    return False


def _render_room_view(meetings: list[Meeting], term_name: str) -> str:
    """One section per room with a schedule grid."""
    groups = _build_meeting_map(meetings)["by_room"]
    slots = _time_slots()

    html = f"<h1>Room Schedule &mdash; {term_name}</h1>\n"

    for room_key in sorted(groups.keys()):
        room_meetings = groups[room_key]
        html += f"<h2>{room_key}</h2>\n"
        html += "<table>\n<tr><th>Time</th>"
        for day in _DAYS:
            html += f"<th>{day}</th>"
        html += "</tr>\n"

        for label, s_start, s_end in slots:
            html += f"<tr><td>{label}</td>"
            for day in _DAYS:
                cell_contents = []
                for m in room_meetings:
                    if _meeting_on_day(m, day) and _meeting_overlaps_slot(m, s_start, s_end):
                        section = m.section
                        course = section.course if section else None
                        code = f"{course.department_code} {course.course_number}" if course else "?"
                        sec = section.section_number if section else ""
                        instr = m.instructor.name if m.instructor else ""
                        cell_contents.append(f"{code}-{sec}<br><small>{instr}</small>")
                if cell_contents:
                    html += f"<td>{'<hr>'.join(cell_contents)}</td>"
                else:
                    html += '<td class="empty"></td>'
            html += "</tr>\n"
        html += "</table>\n"

    return html


def _render_instructor_view(meetings: list[Meeting], term_name: str) -> str:
    """One section per instructor with their schedule."""
    groups = _build_meeting_map(meetings)["by_instructor"]
    slots = _time_slots()

    html = f"<h1>Instructor Schedule &mdash; {term_name}</h1>\n"

    for instr_key in sorted(groups.keys()):
        instr_meetings = groups[instr_key]
        html += f"<h2>{instr_key}</h2>\n"
        html += "<table>\n<tr><th>Time</th>"
        for day in _DAYS:
            html += f"<th>{day}</th>"
        html += "</tr>\n"

        for label, s_start, s_end in slots:
            html += f"<tr><td>{label}</td>"
            for day in _DAYS:
                cell_contents = []
                for m in instr_meetings:
                    if _meeting_on_day(m, day) and _meeting_overlaps_slot(m, s_start, s_end):
                        section = m.section
                        course = section.course if section else None
                        code = f"{course.department_code} {course.course_number}" if course else "?"
                        sec = section.section_number if section else ""
                        room_label = ""
                        if m.room:
                            bldg = m.room.building.abbreviation if m.room.building else ""
                            room_label = f"{bldg} {m.room.room_number}"
                        cell_contents.append(f"{code}-{sec}<br><small>{room_label}</small>")
                if cell_contents:
                    html += f"<td>{'<hr>'.join(cell_contents)}</td>"
                else:
                    html += '<td class="empty"></td>'
            html += "</tr>\n"
        html += "</table>\n"

    return html


def _render_master_view(meetings: list[Meeting], term_name: str) -> str:
    """All rooms as rows, time blocks as columns (master grid)."""
    groups = _build_meeting_map(meetings)["by_room"]
    slots = _time_slots()

    html = f"<h1>Master Schedule &mdash; {term_name}</h1>\n"

    for day in _DAYS:
        html += f"<h2>{day}</h2>\n"
        html += "<table>\n<tr><th>Room</th>"
        for label, _, _ in slots:
            html += f"<th>{label}</th>"
        html += "</tr>\n"

        for room_key in sorted(groups.keys()):
            room_meetings = groups[room_key]
            html += f"<tr><td><strong>{room_key}</strong></td>"
            for _, s_start, s_end in slots:
                cell_contents = []
                for m in room_meetings:
                    if _meeting_on_day(m, day) and _meeting_overlaps_slot(m, s_start, s_end):
                        section = m.section
                        course = section.course if section else None
                        code = f"{course.department_code} {course.course_number}" if course else "?"
                        sec = section.section_number if section else ""
                        instr = m.instructor.name if m.instructor else ""
                        cell_contents.append(f"{code}-{sec}<br><small>{instr}</small>")
                if cell_contents:
                    html += f"<td>{'<hr>'.join(cell_contents)}</td>"
                else:
                    html += '<td class="empty"></td>'
            html += "</tr>\n"
        html += "</table>\n"

    return html


def export_print_html(db: Session, term_id: int, view: str) -> str:
    """Generate printable HTML for the given term and view type.

    Args:
        db: Database session.
        term_id: The term to export.
        view: One of 'room', 'instructor', or 'master'.

    Returns:
        A complete HTML document string.
    """
    term = db.query(Term).filter(Term.id == term_id).first()
    if not term:
        raise ValueError(f"Term {term_id} not found")

    meetings = (
        db.query(Meeting)
        .join(Section, Meeting.section_id == Section.id)
        .filter(Section.term_id == term_id)
        .options(
            joinedload(Meeting.section).joinedload(Section.course),
            joinedload(Meeting.room).joinedload(Room.building),
            joinedload(Meeting.instructor),
        )
        .all()
    )

    if view == "room":
        body = _render_room_view(meetings, term.name)
    elif view == "instructor":
        body = _render_instructor_view(meetings, term.name)
    elif view == "master":
        body = _render_master_view(meetings, term.name)
    else:
        raise ValueError(f"Unknown view type: {view}")

    return f"<!DOCTYPE html>\n<html>\n<head>\n<meta charset=\"utf-8\">\n<title>Schedule &mdash; {term.name}</title>\n{_CSS}</head>\n<body>\n{body}\n</body>\n</html>"
