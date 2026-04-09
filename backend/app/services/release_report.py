"""Export multi-term release planning report (XLSX and HTML)."""
from __future__ import annotations

import io
import os
import re
import sys
from collections import defaultdict
from datetime import datetime
from typing import List, Tuple

from jinja2 import Environment, FileSystemLoader
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.instructor import Instructor
from app.models.load_adjustment import LoadAdjustment
from app.models.term import Term
from app.services.workload import compute_instructor_workload


# Styles (matching workload_export.py)
_header_font = Font(bold=True, size=11)
_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_bold_font = Font(bold=True)
_thin_border = Border(bottom=Side(style="thin"))
_wrap_align = Alignment(wrap_text=True, vertical="top")
_overload_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

_STATUS_MAP = {
    "faculty": "F",
    "ias": "IAS",
    "adjunct": "ADJ",
    "nias": "NIAS",
}

_ADJ_TYPE_LABELS = {
    "research_release": "Research Reassignment",
    "admin_release": "Admin Reassignment",
    "course_release": "Course Reassignment",
    "adhoc": "ADHOC",
    "overload": "Overload",
    "other": "Other",
}


def _gather_release_data(db: Session, term_ids: List[int]) -> dict:
    """Gather all release data across terms for all active instructors.

    Returns dict with keys: terms, instructors, grid.
    grid[instructor_id][term_id] = {adjustments, release_credits, teaching_credits, net_available}
    """
    terms = (
        db.query(Term)
        .filter(Term.id.in_(term_ids))
        .order_by(Term.start_date)
        .all()
    )
    term_map = {t.id: t for t in terms}

    instructors = (
        db.query(Instructor)
        .filter(Instructor.is_active.is_(True))
        .order_by(Instructor.last_name, Instructor.first_name)
        .all()
    )

    # Load adjustments for all requested terms
    adjustments = (
        db.query(LoadAdjustment)
        .filter(LoadAdjustment.term_id.in_(term_ids))
        .all()
    )

    # Group adjustments by (instructor_id, term_id)
    adj_map = defaultdict(list)
    for adj in adjustments:
        adj_map[(adj.instructor_id, adj.term_id)].append(adj)

    # Compute workload per term (for teaching credits context)
    workload_by_term = {}
    for tid in term_ids:
        if tid in term_map:
            workload_by_term[tid] = compute_instructor_workload(db, tid)

    # Build grid
    grid = {}
    for inst in instructors:
        grid[inst.id] = {}
        for t in terms:
            inst_adjs = adj_map.get((inst.id, t.id), [])
            release_credits = sum(a.equivalent_credits for a in inst_adjs)

            # Get teaching credits from workload data
            teaching_credits = 0.0
            wl = workload_by_term.get(t.id)
            if wl:
                for wi in wl["instructors"]:
                    if wi["instructor_id"] == inst.id:
                        teaching_credits = wi["total_teaching_credits"]
                        break

            grid[inst.id][t.id] = {
                "adjustments": [
                    {
                        "description": a.description,
                        "equivalent_credits": a.equivalent_credits,
                        "adjustment_type": a.adjustment_type.value
                        if hasattr(a.adjustment_type, "value")
                        else str(a.adjustment_type),
                    }
                    for a in inst_adjs
                ],
                "release_credits": release_credits,
                "teaching_credits": teaching_credits,
                "net_available": inst.max_credits - release_credits,
            }

    return {
        "terms": terms,
        "instructors": instructors,
        "grid": grid,
    }


def export_release_report_xlsx(
    db: Session, term_ids: List[int]
) -> Tuple[bytes, str]:
    """Generate multi-term release planning XLSX report.

    Returns (xlsx_bytes, suggested_filename).
    """
    data = _gather_release_data(db, term_ids)
    terms = data["terms"]
    instructors = data["instructors"]
    grid = data["grid"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Reassignment Planning"

    # Build headers: Last | First | Type | Max Cr | [per term: Reassignments | Reassign Cr | Teaching Cr | Net Avail] | Total Reassign Cr
    headers = ["Last Name", "First Name", "Type", "Max Credits"]
    for t in terms:
        headers.extend([
            f"{t.name}\nReassignments",
            f"{t.name}\nReassign Cr",
            f"{t.name}\nTeaching Cr",
            f"{t.name}\nNet Available",
        ])
    headers.append("Total Reassign Cr")

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = _wrap_align

    # Column widths
    base_widths = [14, 12, 8, 12]
    term_widths = [24, 12, 12, 12] * len(terms)
    all_widths = base_widths + term_widths + [14]
    for i, w in enumerate(all_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 2
    for inst in instructors:
        status = _STATUS_MAP.get((inst.instructor_type or "").lower(), "")
        ws.cell(row=row, column=1, value=inst.last_name)
        ws.cell(row=row, column=2, value=inst.first_name)
        ws.cell(row=row, column=3, value=status)
        ws.cell(row=row, column=4, value=inst.max_credits)

        col = 5
        total_release = 0.0
        is_overloaded = False

        for t in terms:
            cell_data = grid[inst.id][t.id]
            release_descs = ", ".join(
                f"{a['description']} ({_ADJ_TYPE_LABELS.get(a['adjustment_type'], a['adjustment_type'])})"
                for a in cell_data["adjustments"]
            )
            ws.cell(row=row, column=col, value=release_descs)
            ws.cell(row=row, column=col + 1, value=cell_data["release_credits"])
            ws.cell(row=row, column=col + 2, value=cell_data["teaching_credits"])

            net_cell = ws.cell(row=row, column=col + 3, value=cell_data["net_available"])
            if cell_data["net_available"] < 0:
                net_cell.fill = _overload_fill
                is_overloaded = True

            total_release += cell_data["release_credits"]
            col += 4

        total_cell = ws.cell(row=row, column=col, value=total_release)
        total_cell.font = _bold_font
        if is_overloaded:
            total_cell.fill = _overload_fill

        row += 1

    # Write to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = "reassignment_planning_report.xlsx"
    if len(terms) == 1:
        safe = re.sub(r"[^\w\s-]", "", terms[0].name).strip().replace(" ", "_")
        filename = f"reassignment_planning_{safe}.xlsx"

    return buffer.getvalue(), filename


def export_release_report_html(db: Session, term_ids: List[int]) -> str:
    """Generate multi-term release planning HTML report."""
    data = _gather_release_data(db, term_ids)

    if getattr(sys, "frozen", False):
        templates_dir = os.path.join(sys._MEIPASS, "app", "templates")
    else:
        templates_dir = os.path.join(os.path.dirname(__file__), "..", "templates")

    env = Environment(loader=FileSystemLoader(templates_dir), autoescape=True)
    template = env.get_template("release_report.html")

    # Prepare template data
    instructor_rows = []
    for inst in data["instructors"]:
        status = _STATUS_MAP.get((inst.instructor_type or "").lower(), "")
        term_cells = []
        total_release = 0.0

        for t in data["terms"]:
            cell_data = data["grid"][inst.id][t.id]
            total_release += cell_data["release_credits"]
            term_cells.append({
                "adjustments": cell_data["adjustments"],
                "release_credits": cell_data["release_credits"],
                "teaching_credits": cell_data["teaching_credits"],
                "net_available": cell_data["net_available"],
            })

        instructor_rows.append({
            "last_name": inst.last_name,
            "first_name": inst.first_name,
            "status": status,
            "max_credits": inst.max_credits,
            "term_cells": term_cells,
            "total_release": total_release,
        })

    return template.render(
        terms=[t.name for t in data["terms"]],
        instructors=instructor_rows,
        adj_type_labels=_ADJ_TYPE_LABELS,
        generation_date=datetime.now().strftime("%B %d, %Y at %-I:%M %p"),
    )
