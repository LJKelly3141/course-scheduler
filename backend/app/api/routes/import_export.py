from __future__ import annotations

import io
import json
from datetime import time as dt_time
from typing import Dict, List, Optional, Tuple, Union

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse, HTMLResponse
from sqlalchemy.orm import Session

import openpyxl

from app.database import get_db
from app.models.building import Building
from app.models.room import Room
from app.models.instructor import Instructor, ModalityConstraint
from app.models.course import Course
from app.models.enrollment_record import EnrollmentRecord
from app.models.section import Section, Modality, SectionStatus
from app.models.meeting import Meeting
from app.models.time_block import TimeBlock
from app.models.term import Term
from app.schemas.schemas import (
    ImportPreview,
    ImportResult,
    ScheduleImportPreview,
    ColumnDetectResponse,
    CompareResult,
)
from app.services.export import export_term_csv, export_term_xlsx, export_print_html
from app.services.xlsx_reader import (
    read_xlsx_to_rows,
    read_csv_to_rows,
    ROOM_ALIASES,
    INSTRUCTOR_ALIASES,
    COURSE_ALIASES,
    SCHEDULE_ALIASES,
)
from app.services.xlsx_schedule_parser import (
    read_xlsx_schedule,
    detect_schedule_columns,
    match_time_block,
    find_instructor_matches,
)
from app.services.schedule_compare import compare_schedule

router = APIRouter()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _read_import_file(
    contents: bytes,
    filename: str,
    alias_map: Dict[str, List[str]],
    required_columns: Optional[List[str]] = None,
) -> Tuple[List[dict], List[str]]:
    """Read an uploaded file (XLSX or CSV) and return (rows, messages).

    For XLSX: uses smart column detection via alias_map.
    For CSV: reads as-is (assumes canonical column names from edited preview).
    """
    if filename.lower().endswith(".csv"):
        rows = read_csv_to_rows(contents)
        return rows, []
    else:
        return read_xlsx_to_rows(contents, alias_map, required_columns)


def _validate_columns(rows: list[dict], expected: list[str]) -> list[str]:
    """Check that all expected columns are present. Returns a list of error strings."""
    if not rows:
        return ["File is empty or has no data rows"]
    actual = set(rows[0].keys())
    missing = [col for col in expected if col not in actual]
    if missing:
        return [f"Missing required columns: {', '.join(missing)}"]
    return []


# ---------------------------------------------------------------------------
# POST /import/rooms
# ---------------------------------------------------------------------------
ROOM_COLUMNS = ["building_name", "building_abbreviation", "room_number", "capacity"]


def _validate_room_rows(rows: list[dict]) -> tuple[list[dict], list[str]]:
    """Validate room rows. Returns (valid_rows, errors)."""
    errors = _validate_columns(rows, ROOM_COLUMNS)
    if errors:
        return [], errors

    valid_rows: list[dict] = []
    for i, row in enumerate(rows, start=2):  # row 1 is header
        row_errors: list[str] = []
        building_name = row.get("building_name", "").strip()
        building_abbr = row.get("building_abbreviation", "").strip()
        room_number = row.get("room_number", "").strip()
        capacity_str = row.get("capacity", "").strip()

        if not building_name:
            row_errors.append(f"Row {i}: building_name is required")
        if not building_abbr:
            row_errors.append(f"Row {i}: building_abbreviation is required")
        if not room_number:
            row_errors.append(f"Row {i}: room_number is required")
        if not capacity_str:
            row_errors.append(f"Row {i}: capacity is required")
        else:
            try:
                capacity = int(capacity_str)
                if capacity < 0:
                    row_errors.append(f"Row {i}: capacity must be non-negative")
            except ValueError:
                row_errors.append(f"Row {i}: capacity must be an integer")

        if row_errors:
            errors.extend(row_errors)
        else:
            valid_rows.append({
                "building_name": building_name,
                "building_abbreviation": building_abbr,
                "room_number": room_number,
                "capacity": int(capacity_str),
            })

    return valid_rows, errors


@router.post("/import/rooms", response_model=Union[ImportPreview, ImportResult])
async def import_rooms(
    file: UploadFile = File(...),
    preview: bool = Query(default=True),
    db: Session = Depends(get_db),
):
    contents = await file.read()
    filename = file.filename or ""
    rows, file_messages = _read_import_file(contents, filename, ROOM_ALIASES, ROOM_COLUMNS)
    valid_rows, errors = _validate_room_rows(rows)
    errors = file_messages + errors

    if preview:
        return ImportPreview(
            rows=valid_rows,
            errors=errors,
            valid_count=len(valid_rows),
        )

    # Actual import
    if errors:
        raise HTTPException(status_code=400, detail=errors)

    created = 0
    import_errors: list[str] = []

    for row in valid_rows:
        try:
            # Find or create building
            building = (
                db.query(Building)
                .filter(
                    Building.name == row["building_name"],
                    Building.abbreviation == row["building_abbreviation"],
                )
                .first()
            )
            if not building:
                building = Building(
                    name=row["building_name"],
                    abbreviation=row["building_abbreviation"],
                )
                db.add(building)
                db.flush()

            # Create room
            room = Room(
                building_id=building.id,
                room_number=row["room_number"],
                capacity=row["capacity"],
            )
            db.add(room)
            created += 1
        except Exception as exc:
            import_errors.append(
                f"Error importing room {row['room_number']}: {str(exc)}"
            )

    db.commit()
    return ImportResult(created=created, errors=import_errors)


# ---------------------------------------------------------------------------
# POST /import/instructors
# ---------------------------------------------------------------------------
INSTRUCTOR_COLUMNS = ["name", "email", "department", "modality_constraint", "max_credits"]

VALID_MODALITIES = {e.value for e in ModalityConstraint}


def _validate_instructor_rows(rows: list[dict]) -> tuple[list[dict], list[str]]:
    """Validate instructor rows. Returns (valid_rows, errors)."""
    errors = _validate_columns(rows, INSTRUCTOR_COLUMNS)
    if errors:
        return [], errors

    valid_rows: list[dict] = []
    for i, row in enumerate(rows, start=2):
        row_errors: list[str] = []
        name = row.get("name", "").strip()
        email = row.get("email", "").strip()
        department = row.get("department", "").strip()
        modality = row.get("modality_constraint", "any").strip()
        max_credits_str = row.get("max_credits", "12").strip()

        if not name:
            row_errors.append(f"Row {i}: name is required")
        if not email:
            row_errors.append(f"Row {i}: email is required")
        if not department:
            row_errors.append(f"Row {i}: department is required")
        if modality and modality not in VALID_MODALITIES:
            row_errors.append(
                f"Row {i}: modality_constraint must be one of {', '.join(sorted(VALID_MODALITIES))}"
            )
        if max_credits_str:
            try:
                max_credits = int(max_credits_str)
                if max_credits < 0:
                    row_errors.append(f"Row {i}: max_credits must be non-negative")
            except ValueError:
                row_errors.append(f"Row {i}: max_credits must be an integer")

        if row_errors:
            errors.extend(row_errors)
        else:
            valid_rows.append({
                "name": name,
                "email": email,
                "department": department,
                "modality_constraint": modality or "any",
                "max_credits": int(max_credits_str) if max_credits_str else 12,
            })

    return valid_rows, errors


@router.post("/import/instructors", response_model=Union[ImportPreview, ImportResult])
async def import_instructors(
    file: UploadFile = File(...),
    preview: bool = Query(default=True),
    db: Session = Depends(get_db),
):
    contents = await file.read()
    filename = file.filename or ""
    rows, file_messages = _read_import_file(contents, filename, INSTRUCTOR_ALIASES, INSTRUCTOR_COLUMNS)
    valid_rows, errors = _validate_instructor_rows(rows)
    errors = file_messages + errors

    if preview:
        return ImportPreview(
            rows=valid_rows,
            errors=errors,
            valid_count=len(valid_rows),
        )

    if errors:
        raise HTTPException(status_code=400, detail=errors)

    created = 0
    import_errors: list[str] = []

    for row in valid_rows:
        try:
            # Check for duplicate email
            existing = (
                db.query(Instructor)
                .filter(Instructor.email == row["email"])
                .first()
            )
            if existing:
                import_errors.append(
                    f"Instructor with email {row['email']} already exists (skipped)"
                )
                continue

            instructor = Instructor(
                name=row["name"],
                email=row["email"],
                department=row["department"],
                modality_constraint=ModalityConstraint(row["modality_constraint"]),
                max_credits=row["max_credits"],
            )
            db.add(instructor)
            created += 1
        except Exception as exc:
            import_errors.append(
                f"Error importing instructor {row['name']}: {str(exc)}"
            )

    db.commit()
    return ImportResult(created=created, errors=import_errors)


# ---------------------------------------------------------------------------
# POST /import/courses
# ---------------------------------------------------------------------------
COURSE_COLUMNS = ["department_code", "course_number", "title", "credits"]


def _validate_course_rows(rows: list[dict]) -> tuple[list[dict], list[str]]:
    """Validate course rows. Returns (valid_rows, errors)."""
    errors = _validate_columns(rows, COURSE_COLUMNS)
    if errors:
        return [], errors

    valid_rows: list[dict] = []
    for i, row in enumerate(rows, start=2):
        row_errors: list[str] = []
        dept = row.get("department_code", "").strip()
        number = row.get("course_number", "").strip()
        title = row.get("title", "").strip()
        credits_str = row.get("credits", "").strip()

        if not dept:
            row_errors.append(f"Row {i}: department_code is required")
        if not number:
            row_errors.append(f"Row {i}: course_number is required")
        if not title:
            row_errors.append(f"Row {i}: title is required")
        if not credits_str:
            row_errors.append(f"Row {i}: credits is required")
        else:
            try:
                credits = int(credits_str)
                if credits < 0:
                    row_errors.append(f"Row {i}: credits must be non-negative")
            except ValueError:
                row_errors.append(f"Row {i}: credits must be an integer")

        if row_errors:
            errors.extend(row_errors)
        else:
            valid_rows.append({
                "department_code": dept,
                "course_number": number,
                "title": title,
                "credits": int(credits_str),
            })

    return valid_rows, errors


@router.post("/import/courses", response_model=Union[ImportPreview, ImportResult])
async def import_courses(
    file: UploadFile = File(...),
    preview: bool = Query(default=True),
    db: Session = Depends(get_db),
):
    contents = await file.read()
    filename = file.filename or ""
    rows, file_messages = _read_import_file(contents, filename, COURSE_ALIASES, COURSE_COLUMNS)
    valid_rows, errors = _validate_course_rows(rows)
    errors = file_messages + errors

    if preview:
        return ImportPreview(
            rows=valid_rows,
            errors=errors,
            valid_count=len(valid_rows),
        )

    if errors:
        raise HTTPException(status_code=400, detail=errors)

    created = 0
    import_errors: list[str] = []

    for row in valid_rows:
        try:
            course = Course(
                department_code=row["department_code"],
                course_number=row["course_number"],
                title=row["title"],
                credits=row["credits"],
            )
            db.add(course)
            created += 1
        except Exception as exc:
            import_errors.append(
                f"Error importing course {row['department_code']} {row['course_number']}: {str(exc)}"
            )

    db.commit()
    return ImportResult(created=created, errors=import_errors)


# ---------------------------------------------------------------------------
# POST /import/schedule/detect-columns  (XLSX header detection)
# ---------------------------------------------------------------------------
@router.post("/import/schedule/detect-columns", response_model=ColumnDetectResponse)
async def detect_schedule_columns_endpoint(
    file: UploadFile = File(...),
):
    contents = await file.read()
    file_headers, auto_mapping, warnings = detect_schedule_columns(contents)
    return ColumnDetectResponse(
        file_headers=file_headers,
        column_mapping=auto_mapping,
        canonical_columns=list(SCHEDULE_ALIASES.keys()),
        warnings=warnings,
    )


# ---------------------------------------------------------------------------
# POST /import/schedule  (XLSX)
# ---------------------------------------------------------------------------
@router.post("/import/schedule", response_model=Union[ScheduleImportPreview, ImportResult])
async def import_schedule(
    file: UploadFile = File(...),
    term_id: Optional[int] = Query(default=None),
    preview: bool = Query(default=True),
    instructor_mappings: Optional[str] = Query(default=None),
    column_mapping: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
):
    contents = await file.read()
    filename = file.filename or ""

    # Parse column mapping override if provided
    mapping_override: Optional[dict] = None
    if column_mapping:
        try:
            mapping_override = json.loads(column_mapping)
        except json.JSONDecodeError:
            pass

    # Support CSV (from edited preview) or XLSX (original file)
    if filename.lower().endswith(".csv"):
        csv_rows = read_csv_to_rows(contents)
        valid_rows = []
        errors: list[str] = []
        for row in csv_rows:
            valid_rows.append({k: v for k, v in row.items() if v})
        suggested_term = None
        file_headers: list[str] = []
        detected_mapping: dict = {}
    else:
        valid_rows, errors, suggested_term = read_xlsx_schedule(
            contents, column_mapping_override=mapping_override
        )
        # Also detect columns for the response
        file_headers, detected_mapping, _ = detect_schedule_columns(contents)

    if preview:
        # Collect unique instructor names and find matches
        unique_names = list(dict.fromkeys(
            row["instructor_name"] for row in valid_rows if row.get("instructor_name")
        ))
        existing_instructors = db.query(Instructor).all()
        matches = find_instructor_matches(unique_names, existing_instructors)

        return ScheduleImportPreview(
            rows=valid_rows,
            errors=errors,
            valid_count=len(valid_rows),
            suggested_term=suggested_term,
            instructor_matches=matches,
            file_headers=file_headers,
            column_mapping=mapping_override or detected_mapping,
        )

    if errors:
        raise HTTPException(status_code=400, detail=errors)

    # Validate term exists (required for commit)
    if term_id is None:
        raise HTTPException(status_code=400, detail="term_id is required for import")
    term = db.query(Term).filter(Term.id == term_id).first()
    if not term:
        raise HTTPException(status_code=404, detail=f"Term {term_id} not found")

    # Parse instructor mappings from JSON
    # Format: { "Instructor Name": instructor_id_or_null }
    mappings: dict = {}
    if instructor_mappings:
        try:
            mappings = json.loads(instructor_mappings)
        except json.JSONDecodeError:
            pass

    # Load time blocks for matching
    time_blocks = db.query(TimeBlock).all()

    created = 0
    updated = 0
    import_errors: list[str] = []

    for row in valid_rows:
        try:
            # --- Course: find or create ---
            course = (
                db.query(Course)
                .filter(
                    Course.department_code == row["department_code"],
                    Course.course_number == row["course_number"],
                )
                .first()
            )
            if not course:
                course = Course(
                    department_code=row["department_code"],
                    course_number=row["course_number"],
                    title=row["title"],
                    credits=3,
                )
                db.add(course)
                db.flush()

            # --- Instructor: use mapping or find/create ---
            instructor = None
            if row.get("instructor_name"):
                name = row["instructor_name"]
                mapped_id = mappings.get(name)

                if mapped_id is not None:
                    # User mapped this name to an existing instructor
                    instructor = db.query(Instructor).filter(
                        Instructor.id == int(mapped_id)
                    ).first()
                else:
                    # Try exact match, then create
                    instructor = (
                        db.query(Instructor)
                        .filter(Instructor.name == name)
                        .first()
                    )
                    if not instructor:
                        name_parts = name.lower().split()
                        email_slug = ".".join(name_parts)
                        placeholder_email = f"{email_slug}@uwrf.edu"
                        existing_email = (
                            db.query(Instructor)
                            .filter(Instructor.email == placeholder_email)
                            .first()
                        )
                        if existing_email:
                            placeholder_email = f"{email_slug}.import@uwrf.edu"
                        instructor = Instructor(
                            name=name,
                            email=placeholder_email,
                            department=row["department_code"],
                            modality_constraint=ModalityConstraint.any,
                            max_credits=12,
                        )
                        db.add(instructor)
                        db.flush()

            # --- Building & Room: find or create ---
            room = None
            if row.get("building_name") and row.get("room_number"):
                building = (
                    db.query(Building)
                    .filter(Building.name == row["building_name"])
                    .first()
                )
                if not building:
                    abbr = "".join(
                        w[0].upper() for w in row["building_name"].split() if w
                    )
                    building = Building(
                        name=row["building_name"], abbreviation=abbr
                    )
                    db.add(building)
                    db.flush()
                room = (
                    db.query(Room)
                    .filter(
                        Room.building_id == building.id,
                        Room.room_number == row["room_number"],
                    )
                    .first()
                )
                if not room:
                    room = Room(
                        building_id=building.id,
                        room_number=row["room_number"],
                        capacity=30,
                    )
                    db.add(room)
                    db.flush()

            # --- Modality ---
            days_raw = row.get("days")
            has_schedule = days_raw is not None and days_raw != "" and days_raw != "None"
            modality_raw = row["modality"]
            if modality_raw in ("online_sync", "online_async"):
                modality_val = Modality(modality_raw)
            elif modality_raw == "online":
                modality_val = Modality.online_sync if has_schedule else Modality.online_async
            else:
                modality_val = Modality.in_person
            is_online = modality_val in (Modality.online_sync, Modality.online_async)

            # --- Session ---
            from app.models.section import Session as SessionEnum
            session_raw = row.get("session", "regular")
            try:
                session_val = SessionEnum(session_raw)
            except ValueError:
                session_val = SessionEnum.regular

            # Try to resolve term_session_id from TermSession name lookup
            from app.models.term_session import TermSession
            term_session_id = None
            if session_raw and session_raw != "regular":
                # Map legacy enum names to TermSession names
                _legacy_map = {
                    "session_a": "A", "session_b": "B",
                    "session_c": "C", "session_d": "D",
                }
                ts_name = _legacy_map.get(session_raw, session_raw)
                ts_obj = (
                    db.query(TermSession)
                    .filter(TermSession.term_id == term_id, TermSession.name == ts_name)
                    .first()
                )
                if ts_obj:
                    term_session_id = ts_obj.id

            # --- Section: find or create, update if exists ---
            existing_section = (
                db.query(Section)
                .filter(
                    Section.course_id == course.id,
                    Section.term_id == term_id,
                    Section.section_number == row["section_number"],
                )
                .first()
            )
            is_update = False
            if existing_section:
                # Update existing section with latest import data
                existing_section.modality = modality_val
                existing_section.session = session_val
                existing_section.term_session_id = term_session_id or existing_section.term_session_id
                existing_section.instructor_id = instructor.id if instructor else existing_section.instructor_id
                if has_schedule or is_online:
                    existing_section.status = SectionStatus.scheduled
                section = existing_section
                is_update = True
            else:
                section = Section(
                    course_id=course.id,
                    term_id=term_id,
                    section_number=row["section_number"],
                    enrollment_cap=30,
                    modality=modality_val,
                    session=session_val,
                    term_session_id=term_session_id,
                    status=SectionStatus.scheduled if (has_schedule or is_online) else SectionStatus.unscheduled,
                    instructor_id=instructor.id if instructor else None,
                )
                db.add(section)
            db.flush()

            # --- Meeting: create if days/times exist (update if duplicate) ---
            start_str = row.get("start_time") or ""
            end_str = row.get("end_time") or ""
            if has_schedule and start_str and start_str != "None" and end_str and end_str != "None":
                # days may be a list (from XLSX preview) or a string (from CSV re-upload)
                days_val = row["days"]
                if isinstance(days_val, str):
                    # Try parsing as JSON first (e.g. '["M","W","F"]')
                    try:
                        days_val = json.loads(days_val)
                    except (json.JSONDecodeError, TypeError):
                        # Comma-separated string like "M,W,F"
                        days_val = [d.strip() for d in days_val.split(",") if d.strip()]
                days_json = json.dumps(days_val)
                start = dt_time.fromisoformat(start_str)
                end = dt_time.fromisoformat(end_str)
                tb_id = match_time_block(days_val, start, end, time_blocks)

                # Check for existing meeting with same time/days
                existing_meeting = (
                    db.query(Meeting)
                    .filter(
                        Meeting.section_id == section.id,
                        Meeting.days_of_week == days_json,
                        Meeting.start_time == start,
                        Meeting.end_time == end,
                    )
                    .first()
                )
                if existing_meeting:
                    # Update existing meeting
                    existing_meeting.time_block_id = tb_id
                    existing_meeting.room_id = room.id if room else existing_meeting.room_id
                    existing_meeting.instructor_id = instructor.id if instructor else existing_meeting.instructor_id
                else:
                    meeting = Meeting(
                        section_id=section.id,
                        days_of_week=days_json,
                        start_time=start,
                        end_time=end,
                        time_block_id=tb_id,
                        room_id=room.id if room else None,
                        instructor_id=instructor.id if instructor else None,
                    )
                    db.add(meeting)

            if is_update:
                updated += 1
            else:
                created += 1
        except Exception as exc:
            import_errors.append(
                f"Error importing {row.get('department_code', '?')} "
                f"{row.get('course_number', '?')}-{row.get('section_number', '?')}: {str(exc)}"
            )

    db.commit()
    return ImportResult(created=created, updated=updated, errors=import_errors)


# ---------------------------------------------------------------------------
# POST /terms/{term_id}/compare-schedule
# ---------------------------------------------------------------------------
@router.post("/terms/{term_id}/compare-schedule", response_model=CompareResult)
async def compare_schedule_endpoint(
    term_id: int,
    file: UploadFile = File(...),
    column_mapping: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
):
    contents = await file.read()

    # Parse column mapping override if provided
    mapping_override: Optional[dict] = None
    if column_mapping:
        try:
            mapping_override = json.loads(column_mapping)
        except json.JSONDecodeError:
            pass

    valid_rows, errors, _ = read_xlsx_schedule(
        contents, column_mapping_override=mapping_override
    )

    if errors:
        # Filter out non-blocking warnings; only raise on missing-column errors
        blocking = [e for e in errors if e.startswith("Missing required")]
        if blocking:
            raise HTTPException(status_code=400, detail=blocking)

    try:
        result = compare_schedule(db, term_id, valid_rows)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc))

    return result


# ---------------------------------------------------------------------------
# GET /terms/{term_id}/export/csv
# ---------------------------------------------------------------------------
@router.get("/terms/{term_id}/export/csv")
def export_csv(term_id: int, db: Session = Depends(get_db)):
    try:
        csv_content = export_term_csv(db, term_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc))

    return StreamingResponse(
        iter([csv_content]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=schedule_term_{term_id}.csv"
        },
    )


# ---------------------------------------------------------------------------
# GET /terms/{term_id}/export/xlsx
# ---------------------------------------------------------------------------
@router.get("/terms/{term_id}/export/xlsx")
def export_xlsx(term_id: int, db: Session = Depends(get_db)):
    try:
        xlsx_bytes = export_term_xlsx(db, term_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc))

    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=schedule_term_{term_id}.xlsx"
        },
    )


# ---------------------------------------------------------------------------
# GET /terms/{term_id}/export/print
# ---------------------------------------------------------------------------
@router.get("/terms/{term_id}/export/print")
def export_print(
    term_id: int,
    view: str = Query(default="master", pattern="^(room|instructor|master)$"),
    db: Session = Depends(get_db),
):
    try:
        html_content = export_print_html(db, term_id, view)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc))

    return HTMLResponse(content=html_content)


# ---------------------------------------------------------------------------
# POST /import/enrollment  (multi-sheet XLSX)
# ---------------------------------------------------------------------------
def _derive_semester(start_date_str: str) -> str:
    """Derive semester from start date string like '2024-09-03'."""
    try:
        month = int(str(start_date_str).split("-")[1])
    except (IndexError, ValueError):
        return "Fall"
    if month >= 8:
        return "Fall"
    elif month >= 5:
        return "Summer"
    else:
        return "Spring"


@router.post("/import/enrollment", response_model=Union[ImportPreview, ImportResult])
async def import_enrollment(
    file: UploadFile = File(...),
    preview: bool = Query(default=True),
    db: Session = Depends(get_db),
):
    contents = await file.read()

    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(contents), read_only=True, data_only=True
        )
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not read XLSX: {exc}")

    # Build course lookup for matching
    all_courses = db.query(Course).all()
    course_lookup: dict = {}
    for c in all_courses:
        course_lookup[(c.department_code, c.course_number)] = c.id

    records: list[dict] = []
    errors: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)

        try:
            header = next(rows_iter)
        except StopIteration:
            continue

        columns = [str(c).strip() if c else "" for c in header]

        # Find column indices
        col_map: dict = {}
        expected = {
            "Class Program Code": "dept",
            "Catalog Nbr": "catalog",
            "Class Description": "description",
            "Class Section": "section",
            "Class Type": "class_type",
            "Enrollment Total": "enrolled",
            "Enrollment Max Cap": "cap",
            "Room Capacity": "room_cap",
            "Instruction Mode": "mode",
            "Meeting Pattern": "pattern",
            "Meeting Time Start": "start",
            "Meeting Time End": "end",
            "Instructor Name": "instructor",
            "Minimum Units": "credits",
            "Class Start Date": "start_date",
        }
        for i, col in enumerate(columns):
            if col in expected:
                col_map[expected[col]] = i

        if "dept" not in col_map or "catalog" not in col_map:
            errors.append(f"Sheet {sheet_name}: missing required columns")
            continue

        for row_values in rows_iter:
            if all(v is None for v in row_values):
                continue

            def val(key: str) -> str:
                idx = col_map.get(key)
                if idx is None or idx >= len(row_values):
                    return ""
                v = row_values[idx]
                return str(v).strip() if v is not None else ""

            # Only include enrollment rows
            if val("class_type") != "Enrollment":
                continue

            dept = val("dept")
            catalog = val("catalog")
            if not dept or not catalog:
                continue

            enrolled_str = val("enrolled")
            cap_str = val("cap")
            room_cap_str = val("room_cap")
            credits_str = val("credits")

            try:
                enrolled = int(enrolled_str) if enrolled_str else 0
            except ValueError:
                enrolled = 0
            try:
                cap = int(cap_str) if cap_str else 0
            except ValueError:
                cap = 0
            try:
                room_cap = int(room_cap_str) if room_cap_str else None
            except ValueError:
                room_cap = None
            try:
                credits = int(float(credits_str)) if credits_str else None
            except ValueError:
                credits = None

            semester = _derive_semester(val("start_date"))
            course_id = course_lookup.get((dept, catalog))

            records.append({
                "course_id": course_id,
                "department_code": dept,
                "course_number": catalog,
                "academic_year": sheet_name,
                "semester": semester,
                "section_number": val("section"),
                "enrollment_total": enrolled,
                "enrollment_cap": cap,
                "room_capacity": room_cap,
                "modality": val("mode") or None,
                "meeting_pattern": val("pattern") or None,
                "start_time": val("start") or None,
                "end_time": val("end") or None,
                "instructor_name": val("instructor") or None,
                "credits": credits,
            })

    wb.close()

    if preview:
        matched = sum(1 for r in records if r["course_id"] is not None)
        unmatched = len(records) - matched

        # Build summary rows instead of returning all records
        from collections import Counter
        year_counts = Counter(r["academic_year"] for r in records)
        dept_counts = Counter(r["department_code"] for r in records)
        summary_rows = [
            {
                "stat": "Total enrollment rows",
                "value": str(len(records)),
            },
            {
                "stat": "Matched to scheduled courses",
                "value": str(matched),
            },
            {
                "stat": "Unmatched (other departments/courses)",
                "value": str(unmatched),
            },
            {
                "stat": "Academic years",
                "value": ", ".join(
                    f"{y} ({c})" for y, c in sorted(year_counts.items())
                ),
            },
            {
                "stat": "Departments",
                "value": ", ".join(
                    f"{d} ({c})" for d, c in sorted(dept_counts.items())
                ),
            },
        ]

        # List matched courses
        matched_courses = sorted(set(
            f"{r['department_code']} {r['course_number']}"
            for r in records if r["course_id"] is not None
        ))
        if matched_courses:
            summary_rows.append({
                "stat": "Matched courses",
                "value": ", ".join(matched_courses),
            })

        return ImportPreview(
            rows=summary_rows,
            errors=errors,
            valid_count=len(records),
        )

    # Commit: clear existing records and insert new ones
    db.query(EnrollmentRecord).delete()

    for rec in records:
        db.add(EnrollmentRecord(**rec))

    db.commit()
    matched = sum(1 for r in records if r["course_id"] is not None)
    return ImportResult(
        created=len(records),
        errors=errors + [f"{matched} records matched to existing courses"],
    )
